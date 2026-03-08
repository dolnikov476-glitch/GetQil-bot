import requests
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ─────────────────────────────────────────
# НАСТРОЙКИ
# ─────────────────────────────────────────

KEYWORDS = [
    "электрокарниз",
    "электрокарнизы",
    "электрические карнизы",
    "автоматические карнизы",
    "карнизы с электроприводом",
    "умные шторы",
    "электрошторы",
    "автоматические шторы",
    "шторы с электроприводом",
    "мотор для штор",
]

CITIES = [
    "Москва", "Санкт-Петербург", "Новосибирск", "Екатеринбург",
    "Казань", "Нижний Новгород", "Челябинск", "Самара", "Уфа",
    "Ростов-на-Дону", "Красноярск", "Пермь", "Воронеж", "Краснодар",
    "Саратов", "Тюмень", "Барнаул", "Иркутск", "Омск", "Томск",
    "Хабаровск", "Владивосток", "Ярославль", "Ижевск", "Тольятти",
    "Астрахань", "Пенза", "Липецк", "Тула", "Киров", "Чебоксары",
    "Рязань", "Оренбург", "Набережные Челны", "Ульяновск", "Волгоград",
]

BLACKLIST = [
    "авто", "мотор", "мото", "автосервис", "шиномонтаж",
    "гараж", "стадион", "спортив", "ледов", "тюль", "швейн",
    "автоматизация бизнеса", "автоматизация производства",
    "1с", "бухгалтер", "такси", "транспорт",
]

def is_relevant(name, rubric=""):
    text = (name + " " + rubric).lower()
    for word in BLACKLIST:
        if word in text:
            return False
    return True

# ─────────────────────────────────────────
# ЯНДЕКС СПРАВОЧНИК
# ─────────────────────────────────────────

def search_yandex_sprav(query, city):
    """Парсинг через Яндекс Справочник API"""
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
            "Accept": "application/json, text/javascript, */*",
            "Referer": "https://yandex.ru/",
        }
        r = requests.get(
            "https://yandex.ru/maps/api/search",
            params={
                "text": f"{query} {city}",
                "type": "biz",
                "lang": "ru",
                "results": 20,
                "skip": 0,
            },
            headers=headers,
            timeout=15
        )
        if r.status_code == 200:
            return r.json()
    except Exception as e:
        print(f"    Ошибка: {e}")
    return None

def parse_yandex(all_companies):
    print("\n📋 Парсинг Яндекс Справочника...")
    total = 0

    for keyword in KEYWORDS:
        print(f"\n  🔍 {keyword}")
        kw_found = 0

        for city in CITIES:
            data = search_yandex_sprav(keyword, city)
            if not data:
                time.sleep(0.5)
                continue

            features = []
            if isinstance(data, dict):
                features = (
                    data.get("features", []) or
                    data.get("data", {}).get("features", []) or []
                )

            for feat in features:
                try:
                    props = feat.get("properties", {})
                    name = props.get("name", "").strip()
                    if not name:
                        continue

                    company_meta = props.get("CompanyMetaData", {})
                    rubrics = [c.get("name", "") for c in company_meta.get("Categories", [])]
                    rubric_str = ", ".join(rubrics)

                    if not is_relevant(name, rubric_str):
                        continue

                    # Телефоны
                    phones = []
                    for phone in company_meta.get("Phones", []):
                        num = phone.get("formatted", "") or phone.get("number", "")
                        if num and num not in phones:
                            phones.append(num)

                    # Сайт
                    website = ""
                    for url in company_meta.get("Urls", []):
                        website = url.get("value", "") if isinstance(url, dict) else str(url)
                        if website:
                            break

                    # Адрес
                    address = props.get("description", "")

                    key = f"{name.lower()}_{city.lower()}"
                    if key not in all_companies:
                        all_companies[key] = {
                            "Название": name,
                            "Телефоны": ", ".join(phones),
                            "Сайт": website,
                            "Город": city,
                            "Адрес": address,
                            "Источник": "Яндекс Справочник",
                            "Ключевой запрос": keyword,
                            "Рубрики": rubric_str[:100],
                        }
                        kw_found += 1
                        total += 1
                    else:
                        if phones:
                            existing = all_companies[key]["Телефоны"].split(", ") if all_companies[key]["Телефоны"] else []
                            for p in phones:
                                if p not in existing:
                                    existing.append(p)
                            all_companies[key]["Телефоны"] = ", ".join(existing)
                            all_companies[key]["Источник"] = "2GIS + Яндекс"

                except Exception:
                    continue

            time.sleep(0.35)

        print(f"    Новых: {kw_found}")

    return total

# ─────────────────────────────────────────
# СОХРАНЕНИЕ
# ─────────────────────────────────────────

def save_excel(companies, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Компании"

    HEADER_BG = "1F3864"
    headers = ["№", "Название компании", "Телефон(ы)", "Сайт", "Город", "Адрес", "Источник", "Рубрики"]
    col_widths = [5, 40, 25, 30, 18, 35, 18, 30]

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Электрокарнизы — Яндекс Справочник — {datetime.now().strftime('%d.%m.%Y')}"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="2E5090")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 25

    sorted_companies = sorted(companies, key=lambda x: (0 if x["Телефоны"] else 1, x["Город"]))

    for i, c in enumerate(sorted_companies, 1):
        row = i + 2
        has_phone = bool(c["Телефоны"])
        fill_color = ("E2EFDA" if i % 2 == 0 else "F0F9EC") if has_phone else ("EBF1FA" if i % 2 == 0 else "FFFFFF")
        row_fill = PatternFill("solid", fgColor=fill_color)

        values = [i, c["Название"], c["Телефоны"], c["Сайт"], c["Город"], c["Адрес"], c["Источник"], c["Рубрики"]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            cell.font = Font(name="Arial", size=9)
        ws.row_dimensions[row].height = 18

    total_row = len(sorted_companies) + 3
    ws.merge_cells(f"A{total_row}:H{total_row}")
    with_phone = sum(1 for c in sorted_companies if c["Телефоны"])
    ws[f"A{total_row}"] = f"Всего: {len(sorted_companies)}  |  С телефоном: {with_phone}"
    ws[f"A{total_row}"].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws[f"A{total_row}"].fill = PatternFill("solid", fgColor=HEADER_BG)
    ws[f"A{total_row}"].alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:H{len(sorted_companies)+2}"
    wb.save(filename)

# ─────────────────────────────────────────
# ЗАПУСК
# ─────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 55)
    print("  ПАРСИНГ: ЯНДЕКС СПРАВОЧНИК")
    print("=" * 55)

    all_companies = {}
    parse_yandex(all_companies)
    companies = list(all_companies.values())

    print("\n" + "=" * 55)
    print(f"  ИТОГО: {len(companies)}")
    print(f"  С телефоном: {sum(1 for c in companies if c['Телефоны'])}")
    print("=" * 55)

    if companies:
        filename = f"электрокарнизы_яндекс_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
        save_excel(companies, filename)
        print(f"\n✅ Сохранено: {filename}")
    else:
        print("\n❌ Ничего не найдено — Яндекс заблокировал запросы.")
        print("   Попробуй запустить через VPN!")
