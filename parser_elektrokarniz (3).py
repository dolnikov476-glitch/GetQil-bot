import requests
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# ─────────────────────────────────────────
# КЛЮЧЕВЫЕ СЛОВА
# ─────────────────────────────────────────

KEYWORDS = [
    "электрокарнизы",
    "электрические карнизы",
    "автоматические карнизы",
    "моторизированные карнизы",
    "карнизы с электроприводом",
    "карнизы с мотором",
    "умные карнизы",
    "карнизы на пульте",
    "карнизы с дистанционным управлением",
    "электрические шторы",
    "автоматические шторы",
    "шторы с электроприводом",
    "умные шторы",
    "мотор для штор",
    "автоматизация штор",
    "привод для карниза",
    "электрошторы",
]

BLACKLIST = [
    "гипермаркет", "леруа", "оби", "castorama", "икеа", "ikea",
    "строительный", "стройматериалы", "мебельный", "мебель",
    "тюль", "ткани", "швейная"
]

def is_relevant(name):
    name_lower = name.lower()
    for word in BLACKLIST:
        if word in name_lower:
            return False
    return True

# ─────────────────────────────────────────
# 2GIS
# ─────────────────────────────────────────

def search_2gis(query, page=1):
    try:
        r = requests.get(
            "https://catalog.api.2gis.com/3.0/items",
            params={
                "q": query,
                "country_code": "ru",
                "page_size": 50,
                "page": page,
                "fields": "items.point,items.address,items.contact_groups,items.rubrics",
                "key": "demo",
                "locale": "ru_RU",
            },
            timeout=15
        )
        if r.status_code == 200:
            return r.json()
    except Exception as e:
        print(f"    2GIS ошибка: {e}")
    return None

def extract_phones(contact_groups):
    phones = []
    if not contact_groups:
        return phones
    for group in contact_groups:
        for contact in group.get("contacts", []):
            if contact.get("type") == "phone":
                val = contact.get("value", "")
                if val and val not in phones:
                    phones.append(val)
    return phones

def extract_website(contact_groups):
    if not contact_groups:
        return ""
    for group in contact_groups:
        for contact in group.get("contacts", []):
            if contact.get("type") in ("website", "url"):
                return contact.get("value", "")
    return ""

def parse_2gis(all_companies):
    print("\n📍 Парсинг 2GIS...")
    for i, keyword in enumerate(KEYWORDS, 1):
        print(f"  [{i}/{len(KEYWORDS)}] {keyword}")
        found = 0
        for page in range(1, 4):
            data = search_2gis(keyword, page)
            if not data:
                break
            items = data.get("result", {}).get("items", [])
            if not items:
                break
            for item in items:
                name = item.get("name", "").strip()
                if not name or not is_relevant(name):
                    continue
                address_obj = item.get("address", {})
                city = ""
                if address_obj:
                    for comp in address_obj.get("components", []):
                        if comp.get("type") in ("city", "settlement"):
                            city = comp.get("name", "")
                            break
                    if not city:
                        city = address_obj.get("name", "")
                contact_groups = item.get("contact_groups", [])
                phones = extract_phones(contact_groups)
                website = extract_website(contact_groups)
                rubrics = [r.get("name", "") for r in item.get("rubrics", [])]
                key = f"{name.lower()}_{city.lower()}"
                if key not in all_companies:
                    all_companies[key] = {
                        "Название": name,
                        "Телефоны": ", ".join(phones),
                        "Сайт": website,
                        "Город": city,
                        "Источник": "2GIS",
                        "Ключевой запрос": keyword,
                        "Рубрики": ", ".join(rubrics[:3]),
                    }
                    found += 1
                else:
                    if phones:
                        existing_phones = all_companies[key]["Телефоны"].split(", ") if all_companies[key]["Телефоны"] else []
                        for p in phones:
                            if p not in existing_phones:
                                existing_phones.append(p)
                        all_companies[key]["Телефоны"] = ", ".join(existing_phones)
            time.sleep(0.4)
        print(f"    Новых: {found}")

# ─────────────────────────────────────────
# ЯНДЕКС КАРТЫ (без ключа, через suggest API)
# ─────────────────────────────────────────

CITIES = [
    "Москва", "Санкт-Петербург", "Новосибирск", "Екатеринбург",
    "Казань", "Нижний Новгород", "Челябинск", "Самара", "Уфа",
    "Ростов-на-Дону", "Красноярск", "Пермь", "Воронеж", "Краснодар",
    "Саратов", "Тюмень", "Барнаул", "Иркутск", "Омск", "Томск",
    "Хабаровск", "Владивосток", "Ярославль", "Ижевск", "Тольятти",
]

YANDEX_KEYWORDS = [
    "электрокарнизы",
    "электрические карнизы",
    "автоматические шторы",
    "умные шторы",
]

def search_yandex(query, city):
    """Поиск через Яндекс Карты — неофициальный endpoint без ключа"""
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
            "Referer": "https://yandex.ru/maps/",
            "Accept": "application/json",
        }
        r = requests.get(
            "https://yandex.ru/maps/api/search",
            params={
                "text": f"{query} {city}",
                "type": "biz",
                "lang": "ru",
                "results": 20,
                "origin": "maps-searchbar",
            },
            headers=headers,
            timeout=15
        )
        if r.status_code == 200:
            return r.json()
    except Exception as e:
        print(f"    Яндекс ошибка: {e}")
    return None

def parse_yandex(all_companies):
    print("\n🗺️  Парсинг Яндекс Карт...")
    total_found = 0

    for keyword in YANDEX_KEYWORDS:
        print(f"  Запрос: {keyword}")
        kw_found = 0
        for city in CITIES:
            data = search_yandex(keyword, city)
            if not data:
                time.sleep(0.5)
                continue

            # Яндекс возвращает разные форматы — обрабатываем оба
            features = []
            if isinstance(data, dict):
                features = (
                    data.get("features", []) or
                    data.get("data", {}).get("features", []) or
                    data.get("results", []) or []
                )

            for feat in features:
                try:
                    props = feat.get("properties", {}) if isinstance(feat, dict) else {}
                    name = props.get("name", "").strip()
                    if not name:
                        name = feat.get("name", "").strip()
                    if not name or not is_relevant(name):
                        continue

                    company_meta = props.get("CompanyMetaData", {})

                    # Телефоны
                    phones = []
                    for phone in company_meta.get("Phones", []):
                        num = phone.get("formatted", "") or phone.get("number", "")
                        if num and num not in phones:
                            phones.append(num)

                    # Сайт
                    website = ""
                    for url in company_meta.get("Urls", []):
                        website = url.get("value", "") if isinstance(url, dict) else str(url)
                        if website:
                            break

                    # Рубрики
                    rubrics = [c.get("name", "") for c in company_meta.get("Categories", [])]

                    key = f"{name.lower()}_{city.lower()}"
                    if key not in all_companies:
                        all_companies[key] = {
                            "Название": name,
                            "Телефоны": ", ".join(phones),
                            "Сайт": website,
                            "Город": city,
                            "Источник": "Яндекс Карты",
                            "Ключевой запрос": keyword,
                            "Рубрики": ", ".join(rubrics[:3]),
                        }
                        total_found += 1
                        kw_found += 1
                    else:
                        if phones:
                            existing_phones = all_companies[key]["Телефоны"].split(", ") if all_companies[key]["Телефоны"] else []
                            for p in phones:
                                if p not in existing_phones:
                                    existing_phones.append(p)
                            all_companies[key]["Телефоны"] = ", ".join(existing_phones)
                            all_companies[key]["Источник"] = "2GIS + Яндекс"

                except Exception:
                    continue

            time.sleep(0.4)

        print(f"    Новых: {kw_found}")
    print(f"  Итого из Яндекса: {total_found}")

# ─────────────────────────────────────────
# EXCEL
# ─────────────────────────────────────────

def save_to_excel(companies):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Компании"

    HEADER_BG = "1F3864"
    HEADER_FG = "FFFFFF"
    ROW_ALT   = "EBF1FA"
    ROW_NORM  = "FFFFFF"

    headers = ["№", "Название компании", "Телефон(ы)", "Сайт", "Город/Регион", "Источник", "Ключевой запрос", "Рубрики"]
    col_widths = [5, 40, 25, 30, 20, 15, 30, 30]

    # Заголовок
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Компании по электрокарнизам — 2GIS + Яндекс Карты — {datetime.now().strftime('%d.%m.%Y')}"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Заголовки колонок
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, color=HEADER_FG, size=10)
        cell.fill = PatternFill("solid", fgColor="2E5090")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 25

    # Данные — сначала с телефонами
    companies_sorted = sorted(companies, key=lambda x: (0 if x["Телефоны"] else 1, x["Город"]))

    for i, company in enumerate(companies_sorted, 1):
        row = i + 2
        has_phone = bool(company["Телефоны"])
        if has_phone:
            fill_color = "E2EFDA" if i % 2 == 0 else "F0F9EC"
        else:
            fill_color = ROW_ALT if i % 2 == 0 else ROW_NORM

        row_fill = PatternFill("solid", fgColor=fill_color)
        values = [
            i,
            company["Название"],
            company["Телефоны"],
            company["Сайт"],
            company["Город"],
            company["Источник"],
            company["Ключевой запрос"],
            company["Рубрики"],
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            cell.font = Font(name="Arial", size=9)
        ws.row_dimensions[row].height = 18

    # Итог
    total_row = len(companies_sorted) + 3
    ws.merge_cells(f"A{total_row}:H{total_row}")
    with_phone = sum(1 for c in companies_sorted if c["Телефоны"])
    from_2gis = sum(1 for c in companies_sorted if "2GIS" in c["Источник"])
    from_yandex = sum(1 for c in companies_sorted if "Яндекс" in c["Источник"])
    ws[f"A{total_row}"] = f"Всего: {len(companies_sorted)}  |  С телефоном: {with_phone}  |  2GIS: {from_2gis}  |  Яндекс Карты: {from_yandex}"
    ws[f"A{total_row}"].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws[f"A{total_row}"].fill = PatternFill("solid", fgColor=HEADER_BG)
    ws[f"A{total_row}"].alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:H{len(companies_sorted)+2}"

    filename = f"электрокарнизы_компании_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    wb.save(filename)
    return filename

# ─────────────────────────────────────────
# ЗАПУСК
# ─────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 55)
    print("  ПАРСИНГ: 2GIS + ЯНДЕКС КАРТЫ")
    print("=" * 55)

    all_companies = {}

    parse_2gis(all_companies)
    parse_yandex(all_companies)

    companies = list(all_companies.values())

    print("\n" + "=" * 55)
    print(f"  ИТОГО компаний: {len(companies)}")
    print(f"  С телефоном:    {sum(1 for c in companies if c['Телефоны'])}")
    print(f"  Из 2GIS:        {sum(1 for c in companies if '2GIS' in c['Источник'])}")
    print(f"  Из Яндекса:     {sum(1 for c in companies if 'Яндекс' in c['Источник'])}")
    print("=" * 55)

    if companies:
        filename = save_to_excel(companies)
        print(f"\n  ✅ Сохранено: {filename}")
    else:
        print("\n  ❌ Компании не найдены.")
