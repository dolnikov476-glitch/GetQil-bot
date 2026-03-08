import requests
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Только самые точные запросы по электрокарнизам
KEYWORDS = [
    "электрокарниз",
    "электрокарнизы",
    "электрокарниз купить",
    "электрокарниз установка",
    "электрокарниз монтаж",
    "электрокарниз интернет магазин",
    "электрокарниз оптом",
    "электрокарниз производство",
    "электрокарниз под заказ",
    "электрокарниз цена",
]

BLACKLIST = [
    "гипермаркет", "леруа", "оби", "икеа", "ikea",
    "строительный", "мебельный", "авто", "мотор",
    "тюль", "ткани", "швейная", "спортив", "ледов",
]

def is_relevant(name):
    name_lower = name.lower()
    for word in BLACKLIST:
        if word in name_lower:
            return False
    return True

def search_2gis(query, page=1):
    try:
        r = requests.get(
            "https://catalog.api.2gis.com/3.0/items",
            params={
                "q": query,
                "country_code": "ru",
                "page_size": 50,
                "page": page,
                "fields": "items.point,items.address,items.contact_groups,items.rubrics",
                "key": "demo",
                "locale": "ru_RU",
            },
            timeout=15
        )
        if r.status_code == 200:
            return r.json()
    except Exception as e:
        print(f"    Ошибка: {e}")
    return None

def extract_phones(contact_groups):
    phones = []
    if not contact_groups:
        return phones
    for group in contact_groups:
        for contact in group.get("contacts", []):
            if contact.get("type") == "phone":
                val = contact.get("value", "")
                if val and val not in phones:
                    phones.append(val)
    return phones

def extract_website(contact_groups):
    if not contact_groups:
        return ""
    for group in contact_groups:
        for contact in group.get("contacts", []):
            if contact.get("type") in ("website", "url"):
                return contact.get("value", "")
    return ""

print("=" * 55)
print("  ПАРСИНГ: ЭЛЕКТРОКАРНИЗЫ (расширенный)")
print("=" * 55)

all_companies = {}

for i, keyword in enumerate(KEYWORDS, 1):
    print(f"\n[{i}/{len(KEYWORDS)}] {keyword}")
    found = 0
    for page in range(1, 6):  # до 5 страниц = 250 результатов
        data = search_2gis(keyword, page)
        if not data:
            break
        items = data.get("result", {}).get("items", [])
        if not items:
            break
        for item in items:
            name = item.get("name", "").strip()
            if not name or not is_relevant(name):
                continue
            address_obj = item.get("address", {})
            city = ""
            if address_obj:
                for comp in address_obj.get("components", []):
                    if comp.get("type") in ("city", "settlement"):
                        city = comp.get("name", "")
                        break
                if not city:
                    city = address_obj.get("name", "")
            contact_groups = item.get("contact_groups", [])
            phones = extract_phones(contact_groups)
            website = extract_website(contact_groups)
            rubrics = [r.get("name", "") for r in item.get("rubrics", [])]
            key = f"{name.lower()}_{city.lower()}"
            if key not in all_companies:
                all_companies[key] = {
                    "Название": name,
                    "Телефоны": ", ".join(phones),
                    "Сайт": website,
                    "Город": city,
                    "Источник": "2GIS",
                    "Ключевой запрос": keyword,
                    "Рубрики": ", ".join(rubrics[:3]),
                }
                found += 1
            else:
                if phones:
                    existing = all_companies[key]["Телефоны"].split(", ") if all_companies[key]["Телефоны"] else []
                    for p in phones:
                        if p not in existing:
                            existing.append(p)
                    all_companies[key]["Телефоны"] = ", ".join(existing)
        time.sleep(0.4)
    print(f"    Новых: {found}")

companies = list(all_companies.values())

print("\n" + "=" * 55)
print(f"  ИТОГО: {len(companies)}")
print(f"  С телефоном: {sum(1 for c in companies if c['Телефоны'])}")
print("=" * 55)

# Сохраняем
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Электрокарнизы"

HEADER_BG = "1F3864"
headers = ["№", "Название компании", "Телефон(ы)", "Сайт", "Город/Регион", "Источник", "Ключевой запрос", "Рубрики"]
col_widths = [5, 40, 25, 30, 20, 10, 30, 30]

ws.merge_cells("A1:H1")
ws["A1"] = f"Электрокарнизы — расширенный парсинг — {datetime.now().strftime('%d.%m.%Y')}"
ws["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor=HEADER_BG)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

thin = Side(style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor="2E5090")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
ws.row_dimensions[2].height = 25

companies_sorted = sorted(companies, key=lambda x: (0 if x["Телефоны"] else 1, x["Город"]))

for i, company in enumerate(companies_sorted, 1):
    row = i + 2
    has_phone = bool(company["Телефоны"])
    fill_color = ("E2EFDA" if i % 2 == 0 else "F0F9EC") if has_phone else ("EBF1FA" if i % 2 == 0 else "FFFFFF")
    row_fill = PatternFill("solid", fgColor=fill_color)
    values = [i, company["Название"], company["Телефоны"], company["Сайт"], company["Город"], company["Источник"], company["Ключевой запрос"], company["Рубрики"]]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = row_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border
        cell.font = Font(name="Arial", size=9)
    ws.row_dimensions[row].height = 18

total_row = len(companies_sorted) + 3
ws.merge_cells(f"A{total_row}:H{total_row}")
ws[f"A{total_row}"] = f"Всего: {len(companies_sorted)}  |  С телефоном: {sum(1 for c in companies_sorted if c['Телефоны'])}"
ws[f"A{total_row}"].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
ws[f"A{total_row}"].fill = PatternFill("solid", fgColor=HEADER_BG)
ws[f"A{total_row}"].alignment = Alignment(horizontal="center")

ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:H{len(companies_sorted)+2}"

filename = f"электрокарнизы_расширенный_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
wb.save(filename)
print(f"\n✅ Сохранено: {filename}")
